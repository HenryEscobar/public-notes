import openpyxl

# Play with a spreadsheet. see https://www.youtube.com/watch?v=t8pPdKYpowI for details

if __name__ == '__main__':
    inv_file=openpyxl.load_workbook("inventory.xlsx")
    product_list = inv_file["Sheet1"]

    products_per_supplier = {}
    total_value_per_supplier = {}
    products_under_ten = {}

    for product_row in range(2, product_list.max_row+1):
        supplier_name = product_list.cell(product_row,4).value
        inventory = int(product_list.cell(product_row,2).value)
        price = float(product_list.cell(product_row,3).value)    # cast not needed as default is float
        product_number = int(product_list.cell(product_row,1).value)

        inventory_price = product_list.cell(product_row,5)

        total = inventory * price

        inventory_price.value = total

        if inventory < 10:
            products_under_ten[product_number] = inventory

        # Calculate # of products per supplier
        if supplier_name in products_per_supplier:
            products_per_supplier[supplier_name] += 1
            total_value_per_supplier[supplier_name] += total
        else:
            print("Adding a new suppliers: {}".format(supplier_name))
            products_per_supplier[supplier_name] = 1
            total_value_per_supplier[supplier_name] = total
        

    print(products_per_supplier)
    print(total_value_per_supplier)
    print(products_under_ten)
    inv_file.save("inventory_with_total_value.xlsx")






